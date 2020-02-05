import openpyxl
import datetime
import pandas as pd
from os import sys, path
root_path = path.dirname(path.dirname(path.abspath(__file__)))

try:
    import config
    from Utiltity.common import *
    from Utiltity.time_utility import *
except Exception as e:
    sys.path.append(root_path)

    import config
    from Utiltity.common import *
    from Utiltity.time_utility import *
finally:
    pass


def methods_from_prob(prob: dict) -> []:
    return methods_from_method_list(prob.get('methods', []))


def methods_from_method_list(method_list: list) -> []:
    return [method for method, _, _, _ in method_list]


# def standard_dispatch_analysis(securities: [str], methods: [str], data_hub, database,
#                                method_list: list) -> []:
#     result_list = []
#     for query_method in methods:
#         for hash_id, _, _, function_entry in method_list:
#             if hash_id == query_method:
#                 if function_entry is None:
#                     print('Method ' + hash_id + ' not implemented yet.')
#                 else:
#                     try:
#                         result = function_entry(securities, data_hub, database)
#                     except Exception as e:
#                         print('Execute analyzer [' + hash_id + '] Error: ')
#                         print(e)
#                         print(traceback.format_exc())
#                         result = None
#                     finally:
#                         pass
#                     if result is not None and len(result) > 0:
#                         result_list.append((query_method, result))
#                 break
#     return result_list


# ----------------------------------------------------------------------------------------------------------------------

class AnalysisResult:

    SCORE_MIN = 0
    SCORE_MAX = 100
    SCORE_PASS = SCORE_MAX
    SCORE_FAIL = SCORE_MIN
    SCORE_NOT_APPLIED = None

    def __init__(self, securities: str, score: int or bool, reason: str or [str] = ''):
        self.method = ''
        self.securities = securities

        if isinstance(score, bool):
            self.score = AnalysisResult.SCORE_PASS if score else AnalysisResult.SCORE_FAIL
        elif isinstance(score, (int, float)):
            self.score = score
            self.score = min(self.score, AnalysisResult.SCORE_MAX)
            self.score = max(self.score, AnalysisResult.SCORE_MIN)
        else:
            self.score = score

        if reason is None:
            self.reason = ''
        elif isinstance(reason, (list, tuple)):
            self.reason = '\n'.join(reason)
        else:
            self.reason = reason


# ----------------------------------------------------------------------------------------------------------------------

class AnalysisContext:
    def __init__(self):
        self.cache = {}
        self.extra = {}
        self.logger = None
        self.progress = None


# ----------------------------------------------------------------------------------------------------------------------

def function_entry_example(securities: str, data_hub, database, context: AnalysisContext) -> AnalysisResult:
    """
    The example of analyzer function entry.
    :param securities: A single securities code, should be a str.
    :param data_hub:  DataHubEntry type
    :param database: DatabaseEntry type
    :param context: AnalysisContext type, which can hold cache data for multiple analysis
    :return: AnalysisResult
    """
    pass


method_list_example = [
    ('5c496d06-9961-4157-8d3e-a90683d6d32c', 'analyzer brief', 'analyzer details', function_entry_example),
]


def standard_dispatch_analysis(securities: [str], methods: [str], data_hub, database,
                               extra: dict, method_list: list) -> [(str, [])] or None:
    context = AnalysisContext()
    if isinstance(extra, dict):
        context.extra = extra
        context.logger = extra.get('logger', print)
        context.progress = extra.get('progress', ProgressRate())

    result_list = []
    for query_method in methods:
        sub_list = []
        context.cache.clear()
        for hash_id, _, _, function_entry in method_list:
            if hash_id != query_method:
                continue
            if function_entry is None:
                print('Method ' + hash_id + ' not implemented yet.')
                break
            context.progress.set_progress(hash_id, 0, len(securities))
            for s in securities:
                try:
                    result = function_entry(s, data_hub, database, context)
                except Exception as e:
                    error_info = 'Execute analyzer [' + hash_id + '] for [' + s + '] got exception.'
                    print(error_info)
                    print(e)
                    print(traceback.format_exc())
                    result = AnalysisResult(s, AnalysisResult.SCORE_NOT_APPLIED, error_info)
                finally:
                    context.progress.increase_progress(hash_id)
                    print('Analyzer %s progress: %.2f%%' % (hash_id, context.progress.get_progress_rate(hash_id) * 100))
                if result is None:
                    result = AnalysisResult(s, AnalysisResult.SCORE_NOT_APPLIED, 'NONE')
                sub_list.append(result)

            # Fill result list for alignment
            while len(sub_list) < len(securities):
                sub_list.append(AnalysisResult(securities[len(sub_list)], AnalysisResult.SCORE_NOT_APPLIED, 'NONE'))
            result_list.append((query_method, sub_list))
            context.progress.set_progress(hash_id, len(securities), len(securities))
            break
    return result_list if len(result_list) > 0 else None


# ------------------------------------------------------------------------------------------------------------------


"""
The results should look like:

method1           method2           method3           ...           methodM
m1_result1        m2_result1        m3_result1                      mM_result1
m1_result2        m2_result2        m3_result2                      mM_result2
m1_result3        m2_result3        m3_result3                      mM_result3
.                 .                 .                               .
.                 .                 .                               .
.                 .                 .                               .
m1_resultN        m2_resultN        m3_resultN                      mM_resultN
"""


def get_securities_in_result(result: dict) -> [str]:
    securities = []
    for method, results in result.items():
        for r in results:
            if str_available(r.securities) and r.securities not in securities:
                securities.append(r.securities)
    return securities


def pick_up_pass_securities(result: dict, score_threshold: int, not_applied_as_fail: bool = False) -> [str]:
    securities = get_securities_in_result(result)
    for method, results in result.items():
        for r in results:
            if r.score == AnalysisResult.SCORE_NOT_APPLIED:
                exclude = not_applied_as_fail
            else:
                exclude = (r.score < score_threshold)
            if exclude and r.securities in securities:
                securities.remove(r.securities)
    return securities


# def check_append_report_when_data_missing(df: pd.DataFrame, securities: str,
#                                           uri: str, fields: str or [str], result: list):
#     if df is None or len(df) == 0:
#         error_info = uri + ': Cannot find data for securities : ' + securities
#         log_error(error_info)
#         result.append(AnalysisResult(securities, AnalysisResult.SCORE_NOT_APPLIED, error_info))
#         return True
#     if not isinstance(fields, (list, tuple)):
#         fields = [fields]
#     for field in fields:
#         if field not in df.columns:
#             error_info = uri + ': Field ' + field + ' missing for securities : ' + securities
#             log_error(error_info)
#             result.append(AnalysisResult(securities, AnalysisResult.SCORE_NOT_APPLIED, error_info))
#             return True
#     return False


def check_gen_report_when_data_missing(df: pd.DataFrame, securities: str,
                                       uri: str, fields: str or [str]) -> AnalysisResult or None:
    if df is None or len(df) == 0:
        error_info = uri + ': Cannot find data for securities : ' + securities
        log_error(error_info)
        return AnalysisResult(securities, AnalysisResult.SCORE_NOT_APPLIED, error_info)
    if not isinstance(fields, (list, tuple)):
        fields = [fields]
    for field in fields:
        if field not in df.columns:
            error_info = uri + ': Field ' + field + ' missing for securities : ' + securities
            log_error(error_info)
            return AnalysisResult(securities, AnalysisResult.SCORE_NOT_APPLIED, error_info)
    return None


def gen_report_when_analyzing_error(securities: str, exception: Exception):
    error_info = 'Error when analysing  : ' + securities + '\n'
    error_info += str(exception)
    log_error(error_info)
    print(traceback.format_exc())
    return AnalysisResult(securities, AnalysisResult.SCORE_NOT_APPLIED, error_info)


def generate_analysis_report(result: dict, file_path: str, name_dict: dict = {}):
    wb = openpyxl.Workbook()
    ws_score = wb.active
    ws_score.title = 'Score'
    ws_comments = wb.create_sheet('Comments')

    ws_score['A1'] = 'Securities\\Analyzer'
    ws_comments['A1'] = 'Securities\\Analyzer'

    fill_pass = openpyxl.styles.PatternFill(patternType="solid", start_color="10EF10")
    fill_fail = openpyxl.styles.PatternFill(patternType="solid", start_color="EF1010")
    fill_none = openpyxl.styles.PatternFill(patternType="solid", start_color="808080")

    ROW_OFFSET = 2
    total_score = []

    column = 1
    for analyzer_uuid, analysis_result in result.items():
        # Write securities column
        if column == 1:
            # The first run. Init the total score list here.
            # Flaw: The first column of result should be the full one. Otherwise the index may out of range.
            total_score = [100 for i in range(0, len(analysis_result))]      # 100: Pass; 0: Fail; 50: Pass with None

            row = 2
            col = index_to_excel_column_name(column)
            for r in analysis_result:
                ws_score[col + str(row)] = r.securities
                ws_comments[col + str(row)] = r.securities
                row += 1
            column = 2

        # Write analyzer name
        row = 1
        col = index_to_excel_column_name(column)
        analyzer_name = name_dict.get(analyzer_uuid, analyzer_uuid)
        ws_score[col + str(row)] = analyzer_name
        ws_comments[col + str(row)] = analyzer_name

        # Write scores
        row = ROW_OFFSET
        for r in analysis_result:
            ws_score[col + str(row)] = r.score if r.score is not None else '-'
            ws_comments[col + str(row)] = r.reason
            if r.score is None:
                fill_style = fill_none
                total_score[row - ROW_OFFSET] = 50 if total_score[row - ROW_OFFSET] != 0 else 0
            elif r.score < 50:
                fill_style = fill_fail
                total_score[row - ROW_OFFSET] = 0
            else:
                fill_style = fill_pass
            ws_score[col + str(row)].fill = fill_style
            ws_comments[col + str(row)].fill = fill_style
            row += 1
        column += 1

    # Write total score
    row = 1
    col = index_to_excel_column_name(column)
    for score in total_score:
        if row == 1:
            ws_score[col + str(row)] = 'Total Result'
            row = 2
        if score == 50:
            fill_text = 'PASS'
            fill_style = fill_none
        elif score < 50:
            fill_text = 'FAIL'
            fill_style = fill_fail
        else:
            fill_text = 'PASS'
            fill_style = fill_pass

        ws_score[col + str(row)] = fill_text
        ws_comments[col + str(row)] = fill_text

        ws_score[col + str(row)].fill = fill_style
        ws_comments[col + str(row)].fill = fill_style

        row += 1

    # Write file
    wb.save(file_path)











